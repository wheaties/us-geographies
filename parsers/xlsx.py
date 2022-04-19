from openpyxl import load_workbook


def _sub_none_for_empty(value):
    if isinstance(value, str):
        return value if len(value) else None
    return value


def parse_file(filename, skip_beginning=0, cls=list):
    values = []
    ws = load_workbook(filename, read_only=True, data_only=True).active

    for row in ws.iter_rows(min_row=skip_beginning):
        values.append(cls(_sub_none_for_empty(col.value) for col in row))
    return values
